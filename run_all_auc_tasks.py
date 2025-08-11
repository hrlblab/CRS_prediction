
import pandas as pd
import numpy as np
import os
from sklearn.model_selection import LeaveOneOut
from sklearn.ensemble import RandomForestClassifier
from sklearn.metrics import roc_auc_score
from sklearn.preprocessing import StandardScaler, label_binarize

sheet_map = {'CRSsNP': 'out-s1', 'CRSwNP': 'out-s2'}
feature_sets = {
    'EO': ['a'],
    'CY': ['b'],
    'DM': ['c'],
    'CM': ['d'],
    'MH': ['e'],
    'EO+CY': ['a', 'b'],
    'EO+CY+DM': ['a', 'b', 'c'],
    'EO+CY+DM+CM': ['a', 'b', 'c', 'd'],
    'ALL': ['a', 'b', 'c', 'd', 'e']
}
output_dir = './results_auc'
os.makedirs(output_dir, exist_ok=True)

excel_file = 'C:/hrlblab/CRSproject/yifei_naweed_CRS-project-main/Sep2024respositorydeID2.xlsx'
xl = pd.ExcelFile(excel_file)

from openpyxl import load_workbook
wb = load_workbook(excel_file, data_only=True)

def get_color_groups(sheet_name):
    sheet = wb[sheet_name]
    color_map = {}
    header = next(sheet.iter_rows(min_row=1, max_row=1))
    for i, cell in enumerate(header[3:], start=3):
        color = cell.fill.start_color.rgb
        if color == "FFFF0000":
            group = 'a'
        elif color == "FFAEF9F8":
            group = 'b'
        elif color == "FFFFC000":
            group = 'c'
        elif color == "FFF662AB":
            group = 'd'
        elif color == "FF9CF412":
            group = 'e'
        else:
            group = 'other'
        color_map[i] = group
    return color_map

for label, sheet_name in sheet_map.items():
    df = xl.parse(sheet_name)
    color_group = get_color_groups(sheet_name)
    df_feature_columns = df.columns[3:]
    group_map = {}
    for rel_idx, (abs_idx, g) in enumerate(color_group.items()):
        if g in ['a', 'b', 'c', 'd', 'e']:
            group_map.setdefault(g, []).append(df_feature_columns[rel_idx])
    print(group_map)
    for task_type in ['binary', 'multi']:
        y_col = 'prior_surgery' if task_type == 'binary' else 'number_prior_surgeries'
        df_clean = df.dropna(subset=[y_col]).reset_index(drop=True)

        if task_type == 'multi':
            y = pd.cut(df_clean[y_col], bins=[-1, 0, 1, np.inf], labels=[0, 1, 2])
        else:
            y = df_clean[y_col].astype(int)

        for feature_name, groups in feature_sets.items():
            print(f'=={label}_{feature_name}_{task_type}==')
            selected_features = []
            for g in groups:
                selected_features.extend(group_map.get(g, []))
            X = df_clean[selected_features].copy()

            X = X.apply(lambda col: col.fillna(col.mean()), axis=0)
            X = pd.DataFrame(StandardScaler().fit_transform(X), columns=X.columns)

            auc_scores = []
            loo = LeaveOneOut()

            for seed in range(100):
                model = RandomForestClassifier(random_state=seed)
                print(seed)
                y_true_all = []
                y_prob_all = []

                for train_idx, test_idx in loo.split(X, y):
                    X_train, X_test = X.iloc[train_idx], X.iloc[test_idx]
                    y_train, y_test = y.iloc[train_idx], y.iloc[test_idx]

                    model.fit(X_train, y_train)
                    y_prob = model.predict_proba(X_test)
                    y_true_all.append(y_test.values[0])
                    y_prob_all.append(y_prob[0])

                try:
                    if task_type == 'binary':
                        y_bin = np.vstack((1 - np.array(y_true_all), np.array(y_true_all))).T
                    else:
                        y_bin = label_binarize(y_true_all, classes=[0, 1, 2])
                    auc = roc_auc_score(y_bin, y_prob_all, average='macro', multi_class='ovr')
                    auc_scores.append(auc)
                except Exception as e:
                    print(f"[{label}-{feature_name}-{task_type}] Seed {seed} AUC error: {e}")

            auc_arr = np.array(auc_scores)
            name = f'{label}_{feature_name}_{task_type}'
            np.save(os.path.join(output_dir, f'{name}.npy'), auc_arr)
            pd.DataFrame(auc_arr, columns=["AUC_macro"]).to_csv(os.path.join(output_dir, f'{name}.csv'), index=False)
            print(f"Saved {name}: Mean AUC = {np.mean(auc_arr):.4f}")
